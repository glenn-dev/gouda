"""Deterministic Santander XLSX parsing with an explicit openpyxl boundary."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import Enum
from io import BytesIO
from math import isfinite
from pathlib import Path
import re
from typing import BinaryIO, Iterable, Mapping
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException


PARSER_VERSION = "santander-v0.2"
_DAY_MONTH_RE = re.compile(r"^(\d{1,2})/(\d{1,2})$")
_FULL_DATE_RE = re.compile(r"^\d{1,4}[/-]\d{1,2}[/-]\d{1,4}$")
_COLUMNS = "ABCDEFG"


class RowOutcome(str, Enum):
    PARSED = "PARSED"
    IGNORED = "IGNORED"
    REJECTED = "REJECTED"


class ReconciliationStatus(str, Enum):
    RECONCILED = "RECONCILED"
    NOT_RECONCILED = "NOT_RECONCILED"
    INSUFFICIENT_DATA = "INSUFFICIENT_DATA"
    NOT_APPLICABLE = "NOT_APPLICABLE"


class _SectionState(str, Enum):
    """Small Santander-specific state model for worksheet row interpretation."""

    PRE_MOVEMENT = "pre_movement"
    MOVEMENT_DETAIL = "movement_detail"
    COMMISSION_SUMMARY = "commission_summary"
    POST_SUMMARY = "post_summary"


class ParserError(Exception):
    """Base class for safe workbook-level parser errors."""

    def __init__(self, code: str):
        self.code = code
        super().__init__(code)


class MalformedWorkbookError(ParserError):
    pass


class UnsupportedWorkbookError(ParserError):
    pass


class AmbiguousWorksheetError(ParserError):
    pass


@dataclass(frozen=True, repr=False)
class SourceCell:
    """Neutral cell representation; its value is intentionally not repr-able."""

    value: object | None = field(repr=False)
    cell_type: str | None = None
    number_format: str | None = None
    is_date: bool = False
    is_formula: bool = False

    def __repr__(self) -> str:
        return f"SourceCell(cell_type={self.cell_type!r}, is_date={self.is_date}, is_formula={self.is_formula})"


@dataclass(frozen=True, repr=False)
class RawRecord:
    """Immutable workbook-local interpretation of one source row."""

    raw_record_id: str
    sheet_alias: str
    worksheet_name: str = field(repr=False)
    worksheet_ordinal: int
    row_number: int
    raw_cells: Mapping[str, SourceCell] = field(repr=False)
    row_class: str

    def __repr__(self) -> str:
        return (
            "RawRecord("
            f"sheet_alias={self.sheet_alias!r}, worksheet_ordinal={self.worksheet_ordinal}, "
            f"row_number={self.row_number}, row_class={self.row_class!r})"
        )


@dataclass(frozen=True, repr=False)
class NormalizedMovement:
    source_record_id: str
    occurrence_date: date
    signed_amount: Decimal = field(repr=False)
    currency: str | None = field(default=None, repr=False)
    account_ref: str | None = field(default=None, repr=False)
    description: str | None = field(default=None, repr=False)
    source_reference: str | None = field(default=None, repr=False)
    running_balance: Decimal | None = field(default=None, repr=False)
    provenance: Mapping[str, object] = field(default_factory=dict, repr=False)

    def __repr__(self) -> str:
        return f"NormalizedMovement(source_record_id={self.source_record_id!r}, occurrence_date={self.occurrence_date!r})"


@dataclass(frozen=True, repr=False)
class RowResult:
    outcome: RowOutcome
    raw_record: RawRecord
    movement: NormalizedMovement | None = field(default=None, repr=False)
    error_codes: tuple[str, ...] = ()

    def __repr__(self) -> str:
        return (
            f"RowResult(outcome={self.outcome.value!r}, row_number={self.raw_record.row_number}, "
            f"error_codes={self.error_codes!r})"
        )


@dataclass(frozen=True, repr=False)
class ReconciliationResult:
    status: ReconciliationStatus
    opening_balance: Decimal | None = field(default=None, repr=False)
    ending_balance: Decimal | None = field(default=None, repr=False)
    difference: Decimal | None = field(default=None, repr=False)

    def __repr__(self) -> str:
        return f"ReconciliationResult(status={self.status.value!r})"


@dataclass(frozen=True, repr=False)
class ParseResult:
    sheet_alias: str
    worksheet_name: str = field(repr=False)
    worksheet_ordinal: int
    period_start: date
    period_end: date
    rows: tuple[RowResult, ...]
    reconciliation: ReconciliationResult

    @property
    def parsed_movements(self) -> tuple[NormalizedMovement, ...]:
        return tuple(row.movement for row in self.rows if row.movement is not None)

    @property
    def parsed_count(self) -> int:
        return sum(row.outcome is RowOutcome.PARSED for row in self.rows)

    @property
    def ignored_count(self) -> int:
        return sum(row.outcome is RowOutcome.IGNORED for row in self.rows)

    @property
    def rejected_count(self) -> int:
        return sum(row.outcome is RowOutcome.REJECTED for row in self.rows)

    def __repr__(self) -> str:
        return (
            f"ParseResult(sheet_alias={self.sheet_alias!r}, worksheet_ordinal={self.worksheet_ordinal}, "
            f"parsed={self.parsed_count}, ignored={self.ignored_count}, rejected={self.rejected_count}, "
            f"reconciliation={self.reconciliation.status.value!r})"
        )


@dataclass(frozen=True)
class _SheetSnapshot:
    alias: str
    name: str
    ordinal: int
    visible: bool
    rows: Mapping[int, Mapping[str, SourceCell]]
    max_row: int


@dataclass(frozen=True)
class _Header:
    row_number: int


@dataclass(frozen=True)
class _Period:
    start: date
    end: date


_LABELS = {
    "date": {"date", "fecha", "fechamovimiento", "fechaoperacion"},
    "description": {"description", "descripcion", "detalle", "glosa", "concepto"},
    "debit": {"debit", "cargo", "cargos", "debe"},
    "credit": {"credit", "abono", "abonos", "haber"},
    "balance": {"balance", "saldo"},
}
_PERIOD_START = {"start", "desde", "periodstart", "fechainicio", "startdate"}
_PERIOD_END = {"end", "hasta", "periodend", "fechafin", "enddate"}
# These are sanitized structural labels confirmed in all three source samples.
# Recognition is exact and restricted to standalone marker rows.
_COMMISSION_SUMMARY_MARKER = "resumendecomisiones"
_POST_SUMMARY_MARKER = "mensajes"


def parse_workbook(
    source: Path | str | bytes | BinaryIO,
    *,
    currency: str | None = None,
    account_ref: str | None = None,
) -> ParseResult:
    """Parse one Santander workbook from a path or binary stream.

    The workbook is loaded with ``data_only=False`` so formula-backed cells are
    never silently replaced with cached/calculated values. Formula-backed
    movement rows are rejected and formula-backed period metadata aborts the
    workbook because trusted period derivation is impossible.
    """

    data = _read_source(source)
    try:
        workbook = load_workbook(BytesIO(data), read_only=False, data_only=False, keep_links=False)
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError, IndexError, TypeError) as exc:
        raise MalformedWorkbookError("xlsx_invalid") from exc

    try:
        sheets = [_adapt_sheet(worksheet, ordinal) for ordinal, worksheet in enumerate(workbook.worksheets, 1)]
        candidates = [sheet for sheet in sheets if sheet.visible and _find_header(sheet) is not None]
        if not candidates:
            raise UnsupportedWorkbookError("movement_header_not_found")
        if len(candidates) > 1:
            raise AmbiguousWorksheetError("ambiguous_statement_worksheets")
        return _parse_statement(candidates[0], currency=currency, account_ref=account_ref)
    finally:
        workbook.close()


def _read_source(source: Path | str | bytes | BinaryIO) -> bytes:
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()
    if isinstance(source, bytes):
        return source
    data = source.read()
    if not isinstance(data, bytes):
        raise TypeError("source must provide bytes")
    return data


def _adapt_sheet(worksheet, ordinal: int) -> _SheetSnapshot:
    rows: dict[int, dict[str, SourceCell]] = {}
    max_row = worksheet.max_row or 0
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=worksheet.max_column or 1):
        row_number = row[0].row
        cells: dict[str, SourceCell] = {}
        for cell in row:
            cell_type = getattr(cell, "data_type", None)
            cells[get_column_letter(cell.column)] = SourceCell(
                value=cell.value,
                cell_type=cell_type,
                number_format=getattr(cell, "number_format", None),
                is_date=bool(getattr(cell, "is_date", False)),
                is_formula=cell_type == "f",
            )
        rows[row_number] = cells
    return _SheetSnapshot(
        alias=f"S{ordinal}",
        name=worksheet.title,
        ordinal=ordinal,
        visible=worksheet.sheet_state == "visible",
        rows=rows,
        max_row=max_row,
    )


def _find_header(sheet: _SheetSnapshot) -> _Header | None:
    for row_number in sorted(sheet.rows):
        row = sheet.rows[row_number]
        if any(_column_number(column) > 7 and _present_cell(cell) for column, cell in row.items()):
            continue
        labels = {_column: _normalize_label(cell.value) for _column, cell in row.items() if _present_cell(cell)}
        if (
            _label_matches(labels.get("A"), "date")
            and _label_matches(labels.get("C"), "description")
            and _label_matches(labels.get("E"), "debit")
            and _label_matches(labels.get("F"), "credit")
            and _label_matches(labels.get("G"), "balance")
        ):
            return _Header(row_number)
    return None


def _parse_statement(sheet: _SheetSnapshot, *, currency: str | None, account_ref: str | None) -> ParseResult:
    header = _find_header(sheet)
    if header is None:
        raise UnsupportedWorkbookError("movement_header_not_found")
    period = _extract_period(sheet, header.row_number)
    rows: list[RowResult] = []
    section = _SectionState.PRE_MOVEMENT
    for row_number in range(1, sheet.max_row + 1):
        row_class = "metadata" if row_number < header.row_number else "movement_candidate"
        raw = _raw_record(sheet, row_number, row_class)
        if row_number < header.row_number:
            rows.append(RowResult(RowOutcome.IGNORED, raw, error_codes=("metadata_row",)))
        elif row_number == header.row_number:
            section = _SectionState.MOVEMENT_DETAIL
            rows.append(RowResult(RowOutcome.IGNORED, _replace_class(raw, "header"), error_codes=("header_row",)))
        elif section is _SectionState.MOVEMENT_DETAIL and _is_column_marker(
            sheet.rows.get(row_number, {}), "C", _COMMISSION_SUMMARY_MARKER
        ):
            section = _SectionState.COMMISSION_SUMMARY
            rows.append(_ignored_section_row(raw, "commission_summary_section"))
        elif section is _SectionState.COMMISSION_SUMMARY and _is_column_marker(
            sheet.rows.get(row_number, {}), "A", _POST_SUMMARY_MARKER
        ):
            section = _SectionState.POST_SUMMARY
            rows.append(_ignored_section_row(raw, "post_summary_section"))
        elif section is _SectionState.COMMISSION_SUMMARY:
            rows.append(_ignored_section_row(raw, "commission_summary"))
        elif section is _SectionState.POST_SUMMARY:
            rows.append(_ignored_section_row(raw, "auxiliary_row"))
        elif _looks_like_header(sheet.rows.get(row_number, {})):
            rows.append(RowResult(RowOutcome.IGNORED, _replace_class(raw, "header"), error_codes=("repeated_header",)))
        else:
            rows.append(_interpret_row(raw, period, currency=currency, account_ref=account_ref))
    return ParseResult(
        sheet.alias,
        sheet.name,
        sheet.ordinal,
        period.start,
        period.end,
        tuple(rows),
        _reconcile(rows),
    )


def _is_column_marker(row: Mapping[str, SourceCell], column: str, marker: str) -> bool:
    """Recognize an exact marker in one column with every other cell empty."""

    if _normalize_label(row.get(column, SourceCell(None)).value) != marker:
        return False
    return all(
        not _present_cell(row.get(other_column, SourceCell(None)))
        for other_column in _COLUMNS
        if other_column != column
    )


def _ignored_section_row(raw: RawRecord, reason: str) -> RowResult:
    return RowResult(RowOutcome.IGNORED, _replace_class(raw, "auxiliary"), error_codes=(reason,))


def _raw_record(sheet: _SheetSnapshot, row_number: int, row_class: str) -> RawRecord:
    return RawRecord(
        raw_record_id=f"{sheet.alias}:row:{row_number}",
        sheet_alias=sheet.alias,
        worksheet_name=sheet.name,
        worksheet_ordinal=sheet.ordinal,
        row_number=row_number,
        raw_cells=dict(sheet.rows.get(row_number, {})),
        row_class=row_class,
    )


def _replace_class(raw: RawRecord, row_class: str) -> RawRecord:
    return RawRecord(
        raw.raw_record_id,
        raw.sheet_alias,
        raw.worksheet_name,
        raw.worksheet_ordinal,
        raw.row_number,
        raw.raw_cells,
        row_class,
    )


def _interpret_row(raw: RawRecord, period: _Period, *, currency: str | None, account_ref: str | None) -> RowResult:
    cells = {column: raw.raw_cells.get(column, SourceCell(None)) for column in _COLUMNS}
    if not any(_present_cell(cell) for cell in cells.values()):
        return RowResult(RowOutcome.IGNORED, _replace_class(raw, "blank"), error_codes=("blank_row",))
    if _looks_like_header(raw.raw_cells):
        return RowResult(RowOutcome.IGNORED, _replace_class(raw, "header"), error_codes=("repeated_header",))
    if any(cell.is_formula for cell in cells.values()):
        return RowResult(RowOutcome.REJECTED, raw, error_codes=("formula_unsupported",))

    date_cell = cells["A"]
    debit_cell = cells["E"]
    credit_cell = cells["F"]
    balance_cell = cells["G"]
    date_candidate = _date_candidate(date_cell)
    debit_present = _present_cell(debit_cell)
    credit_present = _present_cell(credit_cell)
    financial_candidate = date_candidate or debit_present or credit_present or _present_cell(balance_cell)
    if not financial_candidate:
        return RowResult(RowOutcome.IGNORED, _replace_class(raw, "auxiliary"), error_codes=("auxiliary_row",))
    if debit_present and credit_present:
        return RowResult(RowOutcome.REJECTED, raw, error_codes=("debit_credit_conflict",))
    if not debit_present and not credit_present:
        return RowResult(RowOutcome.REJECTED, raw, error_codes=("amount_missing",))

    amount_cell = debit_cell if debit_present else credit_cell
    amount, amount_error = _parse_money_cell(amount_cell, role="movement")
    if amount_error is not None:
        return RowResult(RowOutcome.REJECTED, raw, error_codes=(amount_error,))
    occurrence_date, date_error = _interpret_date(date_cell, period)
    if date_error is not None:
        return RowResult(RowOutcome.REJECTED, raw, error_codes=(date_error,))
    assert amount is not None and occurrence_date is not None

    warnings: list[str] = []
    balance = None
    if _present_cell(balance_cell):
        balance, balance_error = _parse_money_cell(balance_cell, role="balance")
        if balance_error is not None:
            warnings.append(balance_error)
    movement = NormalizedMovement(
        source_record_id=raw.raw_record_id,
        occurrence_date=occurrence_date,
        signed_amount=-amount if debit_present else amount,
        currency=currency,
        account_ref=account_ref,
        description=_clean_optional(cells["C"].value),
        source_reference=_clean_optional(cells["D"].value),
        running_balance=balance,
        provenance={
            "sheet_alias": raw.sheet_alias,
            "worksheet_name": raw.worksheet_name,
            "worksheet_ordinal": raw.worksheet_ordinal,
            "row_number": raw.row_number,
            "source_columns": ("A", "E") if debit_present else ("A", "F"),
            "parser_version": PARSER_VERSION,
        },
    )
    return RowResult(RowOutcome.PARSED, raw, movement, tuple(warnings))


def _extract_period(sheet: _SheetSnapshot, header_row: int) -> _Period:
    starts: list[date] = []
    ends: list[date] = []
    for row_number in sorted(sheet.rows):
        if row_number >= header_row:
            break
        row = sheet.rows[row_number]
        for index, column in enumerate(_COLUMNS):
            label_cell = row.get(column, SourceCell(None))
            label = _normalize_label(label_cell.value)
            target = "start" if label in _PERIOD_START else "end" if label in _PERIOD_END else None
            if target is None:
                continue
            for value_column in _COLUMNS[index + 1 :]:
                value_cell = row.get(value_column, SourceCell(None))
                if value_cell.is_formula:
                    raise UnsupportedWorkbookError("formula_unsupported")
                parsed = _parse_full_date_cell(value_cell)
                if parsed is not None:
                    (starts if target == "start" else ends).append(parsed)
                    break
    if len(starts) != 1 or len(ends) != 1:
        raise UnsupportedWorkbookError("period_context_missing" if not starts or not ends else "period_context_ambiguous")
    if starts[0] > ends[0]:
        raise UnsupportedWorkbookError("period_context_invalid")
    return _Period(starts[0], ends[0])


def _reconcile(rows: Iterable[RowResult]) -> ReconciliationResult:
    rows = tuple(rows)
    opening = _metadata_amount(rows, "opening")
    ending = _metadata_amount(rows, "ending")
    parsed = [row.movement for row in rows if row.outcome is RowOutcome.PARSED and row.movement]
    rejected_candidates = [
        row for row in rows
        if row.raw_record.row_class == "movement_candidate" and row.outcome is RowOutcome.REJECTED
    ]
    if not parsed and not rejected_candidates:
        return ReconciliationResult(ReconciliationStatus.NOT_APPLICABLE, opening, ending)
    if rejected_candidates or opening is None or ending is None:
        return ReconciliationResult(ReconciliationStatus.INSUFFICIENT_DATA, opening, ending)
    difference = ending - (opening + sum((movement.signed_amount for movement in parsed), Decimal("0")))
    status = ReconciliationStatus.RECONCILED if difference == 0 else ReconciliationStatus.NOT_RECONCILED
    return ReconciliationResult(status, opening, ending, difference)


def _metadata_amount(rows: Iterable[RowResult], kind: str) -> Decimal | None:
    labels = {"opening": {"openingbalance", "saldoinicial"}, "ending": {"endingbalance", "saldofinal"}}
    for row in rows:
        if row.raw_record.row_class != "metadata":
            continue
        label = _normalize_label(row.raw_record.raw_cells.get("A", SourceCell(None)).value)
        if label in labels[kind]:
            for column in "BCDEFG":
                cell = row.raw_record.raw_cells.get(column, SourceCell(None))
                amount, error = _parse_money_cell(cell, role="balance")
                if error is None and amount is not None:
                    return amount
    return None


def _looks_like_header(row: Mapping[str, SourceCell]) -> bool:
    labels = {_column: _normalize_label(cell.value) for _column, cell in row.items() if _present_cell(cell)}
    return (
        _label_matches(labels.get("A"), "date")
        and _label_matches(labels.get("C"), "description")
        and _label_matches(labels.get("E"), "debit")
        and _label_matches(labels.get("F"), "credit")
        and _label_matches(labels.get("G"), "balance")
    )


def _label_matches(value: str | None, field_name: str) -> bool:
    if not value:
        return False
    if value in _LABELS[field_name]:
        return True
    if field_name == "debit":
        return "cargo" in value or "debit" in value
    if field_name == "credit":
        return "abono" in value or "credit" in value
    return False


def _normalize_label(value: object | None) -> str:
    import unicodedata

    if not isinstance(value, str):
        return ""
    return re.sub(r"[^a-z0-9]", "", unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower())


def _present_cell(cell: SourceCell) -> bool:
    value = cell.value
    if value is None:
        return False
    return not (isinstance(value, str) and value.strip() == "")


def _date_candidate(cell: SourceCell) -> bool:
    if cell.is_formula:
        return True
    if isinstance(cell.value, (date, datetime)):
        return True
    if isinstance(cell.value, (int, float, Decimal)) and not isinstance(cell.value, bool):
        return True
    return isinstance(cell.value, str) and bool(_DAY_MONTH_RE.fullmatch(cell.value.strip()) or _FULL_DATE_RE.fullmatch(cell.value.strip()))


def _parse_full_date_cell(cell: SourceCell) -> date | None:
    value = cell.value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None
    text = value.strip()
    for parts in (text.split("/"), text.split("-")):
        if len(parts) == 3 and all(part.isdigit() for part in parts):
            try:
                if len(parts[0]) == 4:
                    return date(int(parts[0]), int(parts[1]), int(parts[2]))
                return date(int(parts[2]), int(parts[1]), int(parts[0]))
            except ValueError:
                return None
    return None


def _interpret_date(cell: SourceCell, period: _Period) -> tuple[date | None, str | None]:
    if cell.is_formula:
        return None, "formula_unsupported"
    value = cell.value
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return (value, None) if period.start <= value <= period.end else (None, "date_outside_period")
    if not isinstance(value, str):
        return None, "date_unsupported"
    text = value.strip()
    match = _DAY_MONTH_RE.fullmatch(text)
    if match:
        day, month = int(match.group(1)), int(match.group(2))
        if not _calendar_day_month_is_valid(text):
            return None, "date_invalid"
        candidates = []
        for year in range(period.start.year, period.end.year + 1):
            try:
                candidate = date(year, month, day)
            except ValueError:
                continue
            if period.start <= candidate <= period.end:
                candidates.append(candidate)
        return (candidates[0], None) if len(candidates) == 1 else (None, "date_year_ambiguous")
    if _FULL_DATE_RE.fullmatch(text):
        parsed = _parse_full_date_cell(SourceCell(text))
        return (parsed, None) if parsed is not None and period.start <= parsed <= period.end else (None, "date_invalid" if parsed is None else "date_outside_period")
    return None, "date_invalid"


def _calendar_day_month_is_valid(value: str) -> bool:
    match = _DAY_MONTH_RE.fullmatch(value.strip())
    if match is None:
        return False
    try:
        date(2000, int(match.group(2)), int(match.group(1)))
    except ValueError:
        return False
    return True


def _parse_money_cell(cell: SourceCell, *, role: str) -> tuple[Decimal | None, str | None]:
    if not _present_cell(cell):
        return None, None
    if cell.is_formula:
        return None, "formula_unsupported"
    value = cell.value
    if isinstance(value, bool) or isinstance(value, (date, datetime)):
        return None, "amount_invalid"
    if isinstance(value, (int, float, Decimal)):
        try:
            if isinstance(value, float) and not isfinite(value):
                return None, "amount_invalid"
            amount = Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None, "amount_invalid"
    elif isinstance(value, str):
        normalized = value.strip().replace("\u00a0", "").replace(" ", "").replace("$", "")
        if normalized.startswith("-") or normalized.startswith("("):
            return None, "negative_source_amount"
        amount = _parse_money_text(value)
        if amount is None:
            return None, "amount_invalid"
    else:
        return None, "amount_invalid"
    if amount < 0:
        return None, "negative_source_amount"
    if amount == 0 and role == "movement":
        return None, "zero_amount_unsupported"
    return amount, None


def _parse_money_text(value: str) -> Decimal | None:
    text = value.strip().replace("\u00a0", "").replace(" ", "")
    if not text or text.startswith("-") or text.startswith("+") or text.startswith("(") or text.endswith(")"):
        return None
    if "$" in text:
        if not text.startswith("$"):
            return None
        text = text[1:]
    if not re.fullmatch(r"[0-9]+(?:[.,][0-9]+)*", text):
        return None
    separators = set(char for char in text if char in ".,")
    if len(separators) == 2:
        decimal_separator = "," if text.rfind(",") > text.rfind(".") else "."
        grouping_separator = "." if decimal_separator == "," else ","
        integer, fractional = text.rsplit(decimal_separator, 1)
        groups = integer.split(grouping_separator)
        if len(fractional) not in (1, 2) or len(groups) < 2 or any(len(group) != 3 for group in groups[1:]):
            return None
        text = "".join(groups) + "." + fractional
    elif len(separators) == 1:
        separator = next(iter(separators))
        parts = text.split(separator)
        if len(parts) > 2:
            if any(len(group) != 3 for group in parts[1:]):
                return None
            text = "".join(parts)
        else:
            fractional_length = len(parts[1])
            if fractional_length in (1, 2):
                text = f"{parts[0]}.{parts[1]}"
            elif fractional_length == 3:
                text = "".join(parts)
            else:
                return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _clean_optional(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, str) and value.strip():
        return value.strip()
    return None


def _column_number(column: str) -> int:
    result = 0
    for char in column:
        if not ("A" <= char <= "Z"):
            return 0
        result = result * 26 + ord(char) - ord("A") + 1
    return result
