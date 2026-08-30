from __future__ import annotations

from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook


METADATA = (
    "Saldo Disponible",
    "Saldo Contable",
    "Retenciones",
    "Sobregiro Disponible",
    "Sobregiro Utilizado",
    "Línea de Emergencia",
)
HEADERS = {
    "A8": "Fecha Transacción",
    "B8": "Fecha Contable",
    "C8": "Descripción",
    "G8": "Cargo $",
    "H8": "Abono $",
}


def synthetic_recent_movements_xlsx(
    *,
    rows: list[dict[str, object | None]] | None = None,
    declared_dimension: str | None = None,
    sheet_name: str = "movimientos",
    include_merges: bool = True,
    header_overrides: dict[str, object | None] | None = None,
    metadata_overrides: dict[int, object | None] | None = None,
    extra_sheet: bool = False,
) -> bytes:
    rows = rows or [
        {
            "transaction_date": "25/08/2026",
            "accounting_date": "25/08/2026",
            "description": "Synthetic cargo",
            "cargo": "1.234",
        },
        {
            "transaction_date": "24/08/2026",
            "accounting_date": "25/08/2026",
            "description": "Synthetic abono",
            "abono": "2.345",
        },
    ]
    workbook = Workbook()
    sheet = workbook.active
    _populate_sheet(
        sheet,
        rows=rows,
        sheet_name=sheet_name,
        include_merges=include_merges,
        header_overrides=header_overrides,
        metadata_overrides=metadata_overrides,
    )
    if extra_sheet:
        duplicate = workbook.create_sheet("movimientos-copy")
        _populate_sheet(duplicate, rows=rows, sheet_name="movimientos", include_merges=include_merges)

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    data = stream.getvalue()
    if declared_dimension is None:
        return data
    return _replace_declared_dimension(data, declared_dimension)


def _populate_sheet(
    sheet,
    *,
    rows: list[dict[str, object | None]],
    sheet_name: str,
    include_merges: bool,
    header_overrides: dict[str, object | None] | None = None,
    metadata_overrides: dict[int, object | None] | None = None,
) -> None:
    sheet.title = sheet_name
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "Últimos Movimientos"
    for index, label in enumerate(METADATA, 2):
        sheet.cell(index, 4).value = (metadata_overrides or {}).get(index, label)
        sheet.cell(index, 5).value = "Synthetic snapshot"
    for coordinate, value in {**HEADERS, **(header_overrides or {})}.items():
        sheet[coordinate] = value
    if include_merges:
        sheet.merge_cells("C8:F8")
    for row_number, row in enumerate(rows, 9):
        if include_merges:
            sheet.merge_cells(start_row=row_number, start_column=3, end_row=row_number, end_column=6)
        sheet.cell(row_number, 1).value = row.get("transaction_date")
        sheet.cell(row_number, 2).value = row.get("accounting_date")
        sheet.cell(row_number, 3).value = row.get("description")
        sheet.cell(row_number, 7).value = row.get("cargo")
        sheet.cell(row_number, 8).value = row.get("abono")


def _replace_declared_dimension(data: bytes, dimension: str) -> bytes:
    with ZipFile(BytesIO(data)) as source:
        members = {name: source.read(name) for name in source.namelist()}
    sheet_name = "xl/worksheets/sheet1.xml"
    xml = members[sheet_name]
    marker = b"<dimension ref=\""
    start = xml.index(marker) + len(marker)
    end = xml.index(b"\"", start)
    members[sheet_name] = xml[:start] + dimension.encode("ascii") + xml[end:]
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as target:
        for name, content in members.items():
            target.writestr(name, content)
    return output.getvalue()
