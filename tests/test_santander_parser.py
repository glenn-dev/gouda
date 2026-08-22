from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import unittest

from openpyxl import Workbook

from gouda.santander_parser import (
    AmbiguousWorksheetError,
    MalformedWorkbookError,
    ReconciliationStatus,
    RowOutcome,
    UnsupportedWorkbookError,
    parse_workbook,
)


FIXTURE = Path(__file__).parent / "fixtures" / "santander" / "synthetic-current-account.xlsx"


class SantanderParserTests(unittest.TestCase):
    def test_fixture_recognizes_observed_layout_and_explicit_outcomes(self):
        result = parse_workbook(FIXTURE, currency="ZZZ", account_ref="fixture-account")

        self.assertEqual((result.period_start, result.period_end), (date(2026, 11, 2), date(2026, 11, 30)))
        self.assertEqual(result.parsed_count, 3)
        self.assertEqual(result.rejected_count, 1)
        self.assertGreaterEqual(result.ignored_count, 1)
        self.assertEqual(result.reconciliation.status, ReconciliationStatus.INSUFFICIENT_DATA)
        self.assertTrue(all(row.outcome in (RowOutcome.PARSED, RowOutcome.IGNORED, RowOutcome.REJECTED) for row in result.rows))

    def test_positive_debit_is_negative_and_positive_credit_is_positive(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["04/02", "cargo", "Debit", "REF-1", "$1.00", None, "$9.00"],
            ["05/02", "abono", "Credit", "REF-2", None, "$2.00", "$11.00"],
        ]))

        self.assertEqual([movement.signed_amount for movement in result.parsed_movements], [Decimal("-1.00"), Decimal("2.00")])

    def test_negative_debit_and_credit_are_rejected_without_sign_inversion(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["04/02", "cargo", "Negative debit", "REF-NEG-D", "-1.00", None, "$9.00"],
            ["05/02", "abono", "Negative credit", "REF-NEG-C", None, "-2.00", "$7.00"],
        ]))

        rejected = [row for row in result.rows if row.outcome is RowOutcome.REJECTED]
        self.assertEqual([row.error_codes for row in rejected], [("negative_source_amount",), ("negative_source_amount",)])
        self.assertEqual(result.parsed_count, 0)

    def test_debit_and_credit_conflict_is_rejected(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["04/02", "cargo", "Conflict", "REF-CONFLICT", "$1", "$2", "$9"],
        ]))

        rejected = [row for row in result.rows if row.outcome is RowOutcome.REJECTED]
        self.assertEqual(rejected[0].error_codes, ("debit_credit_conflict",))

    def test_zero_movement_amount_is_rejected_by_explicit_policy(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["04/02", "cargo", "Zero", "REF-ZERO", 0, None, 10],
        ]))

        rejected = [row for row in result.rows if row.outcome is RowOutcome.REJECTED]
        self.assertEqual(rejected[0].error_codes, ("zero_amount_unsupported",))

    def test_decimal_amount_is_exact_for_numeric_and_text_cells(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["03/02", "cargo", "Numeric decimal", "REF-DEC-N", Decimal("1.10"), None, Decimal("8.90")],
            ["04/02", "abono", "Text decimal", "REF-DEC-T", None, "$2.20", "$11.10"],
        ], opening="$10.00", ending="$11.10"))

        self.assertEqual([movement.signed_amount for movement in result.parsed_movements], [Decimal("-1.10"), Decimal("2.20")])
        self.assertEqual(result.reconciliation.status, ReconciliationStatus.RECONCILED)

    def test_rejected_movement_makes_reconciliation_insufficient(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["04/02", "cargo", "Valid", "REF-1", "$1", None, "$9"],
            ["05/02", "cargo", "Ambiguous", "REF-2", "$1", "$1", "$9"],
        ], opening="$10", ending="$9"))

        self.assertEqual(result.parsed_count, 1)
        self.assertEqual(result.reconciliation.status, ReconciliationStatus.INSUFFICIENT_DATA)

    def test_complete_arithmetic_mismatch_is_not_reconciled(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["04/02", "cargo", "Mismatch", "REF-1", "$1", None, "$9"],
        ], opening="$10", ending="$12"))

        self.assertEqual(result.reconciliation.status, ReconciliationStatus.NOT_RECONCILED)

    def test_missing_reconciliation_evidence_does_not_invalidate_parsing(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["04/02", "cargo", "No balance metadata", "REF-1", "$1", None, None],
        ], opening=None, ending=None))

        self.assertEqual(result.parsed_count, 1)
        self.assertEqual(result.reconciliation.status, ReconciliationStatus.INSUFFICIENT_DATA)

    def test_nonfinancial_only_workbook_is_not_applicable(self):
        result = parse_workbook(workbook_bytes(rows=[["Nota auxiliar", None, None, None, None, None, None]], opening="$10", ending="$10"))

        self.assertEqual(result.parsed_count, 0)
        self.assertEqual(result.reconciliation.status, ReconciliationStatus.NOT_APPLICABLE)

    def test_missing_direction_on_date_row_is_rejected(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["06/02", "cargo", "Missing amount", "REF-MISS", None, None, "$10"],
        ]))

        rejected = [row for row in result.rows if row.outcome is RowOutcome.REJECTED]
        self.assertEqual(rejected[0].error_codes, ("amount_missing",))

    def test_unsupported_description_column_fails_closed(self):
        result = workbook_bytes(description_column="B")

        with self.assertRaises(UnsupportedWorkbookError) as context:
            parse_workbook(result)
        self.assertEqual(context.exception.code, "movement_header_not_found")

    def test_provenance_retains_name_ordinal_and_row(self):
        result = parse_workbook(workbook_bytes(sheet_name="Synthetic Audit Sheet", rows=[
            ["04/02", "cargo", "Synthetic description", "SYN-REF", "$1", None, "$9"],
        ]))
        movement = result.parsed_movements[0]

        self.assertEqual(result.worksheet_name, "Synthetic Audit Sheet")
        self.assertEqual(result.worksheet_ordinal, 1)
        self.assertEqual(movement.provenance["worksheet_name"], "Synthetic Audit Sheet")
        self.assertEqual(movement.provenance["worksheet_ordinal"], 1)
        self.assertEqual(movement.provenance["row_number"], 22)
        self.assertEqual(movement.source_record_id, "S1:row:22")

    def test_sensitive_values_are_absent_from_representations(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["04/02", "cargo", "SECRET_DESCRIPTION_908172", "SECRET_REFERENCE_908172", "$1", None, "$9"],
        ], opening="$10", ending="$9"))
        rendered = " ".join(repr(value) for value in (result, result.rows[21], result.parsed_movements[0], result.reconciliation))

        self.assertNotIn("SECRET_DESCRIPTION_908172", rendered)
        self.assertNotIn("SECRET_REFERENCE_908172", rendered)
        self.assertNotIn("$1", rendered)
        self.assertNotIn("$10", rendered)

    def test_repeated_header_is_ignored(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["04/02", "cargo", "Valid", "REF-1", "$1", None, "$9"],
            ["Date", "Type", "Description", "Reference", "Debit", "Credit", "Balance"],
        ]))

        repeated = [row for row in result.rows if row.error_codes == ("repeated_header",)]
        self.assertEqual(len(repeated), 1)
        self.assertEqual(repeated[0].outcome, RowOutcome.IGNORED)

    def test_december_to_january_derives_year(self):
        result = parse_workbook(workbook_bytes(
            period_start="15/12/2025",
            period_end="14/01/2026",
            rows=[
                ["29/12", "cargo", "December", "REF-DEC", "$1", None, "$9"],
                ["03/01", "abono", "January", "REF-JAN", None, "$2", "$11"],
            ],
            opening="$10",
            ending="$11",
        ))

        self.assertEqual([movement.occurrence_date for movement in result.parsed_movements], [date(2025, 12, 29), date(2026, 1, 3)])

    def test_leap_day_is_accepted_inside_leap_period(self):
        result = parse_workbook(workbook_bytes(
            period_start="28/02/2024",
            period_end="29/02/2024",
            rows=[["29/02", "cargo", "Leap day", "REF-LEAP", "$1", None, "$9"]],
        ))

        self.assertEqual(result.parsed_movements[0].occurrence_date, date(2024, 2, 29))

    def test_invalid_day_month_is_rejected(self):
        result = parse_workbook(workbook_bytes(rows=[
            ["31/02", "cargo", "Invalid date", "REF-DATE", "$1", None, "$9"],
        ]))

        rejected = [row for row in result.rows if row.outcome is RowOutcome.REJECTED]
        self.assertEqual(rejected[0].error_codes, ("date_invalid",))

    def test_ambiguous_long_period_is_rejected(self):
        result = parse_workbook(workbook_bytes(
            period_start="01/01/2025",
            period_end="31/12/2026",
            rows=[["03/01", "cargo", "Ambiguous year", "REF-YEAR", "$1", None, "$9"]],
        ))

        rejected = [row for row in result.rows if row.outcome is RowOutcome.REJECTED]
        self.assertEqual(rejected[0].error_codes, ("date_year_ambiguous",))

    def test_full_excel_datetime_must_fall_inside_period(self):
        result = parse_workbook(workbook_bytes(rows=[[datetime(2026, 2, 4), "cargo", "Datetime", "REF-DT", "$1", None, "$9"]]))

        self.assertEqual(result.parsed_movements[0].occurrence_date, date(2026, 2, 4))

    def test_excel_datetime_outside_period_is_rejected(self):
        result = parse_workbook(workbook_bytes(rows=[[datetime(2027, 2, 4), "cargo", "Datetime", "REF-DT", "$1", None, "$9"]]))

        rejected = [row for row in result.rows if row.outcome is RowOutcome.REJECTED]
        self.assertEqual(rejected[0].error_codes, ("date_outside_period",))

    def test_undecoded_excel_date_serial_is_rejected(self):
        result = parse_workbook(workbook_bytes(rows=[[45292, "cargo", "Serial date", "REF-SERIAL", "$1", None, "$9"]]))

        rejected = [row for row in result.rows if row.outcome is RowOutcome.REJECTED]
        self.assertEqual(rejected[0].error_codes, ("date_unsupported",))

    def test_formula_financial_cell_is_rejected_without_using_cached_value(self):
        result = parse_workbook(workbook_bytes(formula_amount=True))

        rejected = [row for row in result.rows if row.outcome is RowOutcome.REJECTED]
        self.assertEqual(rejected[0].error_codes, ("formula_unsupported",))

    def test_formula_period_metadata_is_workbook_failure(self):
        with self.assertRaises(UnsupportedWorkbookError) as context:
            parse_workbook(workbook_bytes(formula_period=True))
        self.assertEqual(context.exception.code, "formula_unsupported")

    def test_hidden_statement_sheet_is_not_selected(self):
        workbook = workbook_bytes(header=False, hidden_statement=True)

        with self.assertRaises(UnsupportedWorkbookError) as context:
            parse_workbook(workbook)
        self.assertEqual(context.exception.code, "movement_header_not_found")

    def test_no_candidate_sheet_is_explicit_failure(self):
        with self.assertRaises(UnsupportedWorkbookError) as context:
            parse_workbook(workbook_bytes(header=False))
        self.assertEqual(context.exception.code, "movement_header_not_found")

    def test_multiple_candidate_sheets_are_ambiguous(self):
        with self.assertRaises(AmbiguousWorksheetError) as context:
            parse_workbook(workbook_bytes(extra_statement=True))
        self.assertEqual(context.exception.code, "ambiguous_statement_worksheets")

    def test_malformed_package_is_typed(self):
        with self.assertRaises(MalformedWorkbookError) as context:
            parse_workbook(BytesIO(b"not-an-xlsx"))
        self.assertEqual(context.exception.code, "xlsx_invalid")

    def test_commission_summary_financial_row_is_ignored_by_section_contract(self):
        result = parse_workbook(workbook_bytes(
            opening="$100",
            ending="$102",
            rows=[
                ["04/02", "cargo", "Primary synthetic debit", "PRIMARY-D", "$3", None, "$97"],
                ["05/02", "abono", "Primary synthetic credit", "PRIMARY-C", None, "$5", "$102"],
                [None, None, "Resumen de Comisiones", None, None, None, None],
                ["04/02", "cargo", "Summary synthetic debit", "SUMMARY-D", "$3", None, "$97"],
                ["MENSAJES", None, None, None, None, None, None],
            ],
        ))

        summary = next(row for row in result.rows if row.raw_record.row_number == 25)
        start_marker = next(row for row in result.rows if row.raw_record.row_number == 24)
        post_summary_marker = next(row for row in result.rows if row.raw_record.row_number == 26)

        self.assertEqual(result.parsed_count, 2)
        self.assertEqual(start_marker.outcome, RowOutcome.IGNORED)
        self.assertEqual(post_summary_marker.outcome, RowOutcome.IGNORED)
        self.assertEqual(summary.outcome, RowOutcome.IGNORED)
        self.assertEqual(summary.error_codes, ("commission_summary",))
        self.assertIsNone(summary.movement)
        self.assertEqual(result.reconciliation.status, ReconciliationStatus.RECONCILED)


def workbook_bytes(
    *,
    period_start: str = "01/02/2026",
    period_end: str = "28/02/2026",
    opening: str | None = "$10",
    ending: str | None = "$10",
    rows: list[list[object]] | None = None,
    header: bool = True,
    description_column: str = "C",
    sheet_name: str = "Synthetic statement",
    hidden_statement: bool = False,
    extra_statement: bool = False,
    formula_amount: bool = False,
    formula_period: bool = False,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    populate_sheet(sheet, period_start, period_end, opening, ending, rows, header, description_column, formula_amount, formula_period)
    if hidden_statement:
        sheet.sheet_state = "hidden"
        visible = workbook.create_sheet("Not a statement")
        visible["A1"] = "Unrelated"
    if extra_statement:
        second = workbook.create_sheet("Second statement")
        populate_sheet(second, period_start, period_end, opening, ending, rows, header, description_column, formula_amount, formula_period)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def populate_sheet(sheet, period_start, period_end, opening, ending, rows, header, description_column, formula_amount, formula_period):
    sheet["A1"] = "Institution"
    sheet["B1"] = "FixtureOnly Bank"
    sheet["E1"] = "Account"
    sheet["A5"] = "Start"
    sheet["B5"] = "=DATE(2026,2,1)" if formula_period else period_start
    sheet["E5"] = "End"
    sheet["F5"] = period_end
    sheet["A9"] = "Account identifier"
    sheet["B9"] = "FIXTURE-ACCOUNT"
    sheet["A11"] = "Opening balance"
    sheet["B11"] = opening
    sheet["A12"] = "Ending balance"
    sheet["B12"] = ending
    sheet["A20"] = "Account movements"
    if not header:
        return
    headers = {"A": "Date", "B": "Type", "C": "Description", "D": "Reference", "E": "Cargos del periodo", "F": "Abonos del periodo", "G": "Balance"}
    if description_column == "B":
        headers["B"], headers["C"] = "Description", "Type"
    for column, value in headers.items():
        sheet[f"{column}21"] = value
    rows = rows or [["04/02", "cargo", "Fixture movement", "REF-1", "$1", None, "$9"]]
    for row_number, values in enumerate(rows, 22):
        for column, value in zip("ABCDEFG", values):
            sheet[f"{column}{row_number}"] = value
    if formula_amount:
        sheet["E22"] = "=1+1"
